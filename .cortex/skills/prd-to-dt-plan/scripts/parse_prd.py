"""Parse XLSX or CSV PRD files into JSON for the prd-to-dt-plan skill.

Usage:
    uv run --project <SKILL_DIR> python <SKILL_DIR>/scripts/parse_prd.py file1.xlsx [file2.csv ...]

Output:
    JSON object to stdout. Each key is "<filename>::<sheet>" (XLSX) or "<filename>" (CSV).
    Each value is a list of row dicts.
"""

import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_xlsx(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            data.append({h: (str(v).strip() if v is not None else None) for h, v in zip(headers, row)})
        result[f"{path.name}::{sheet_name}"] = data
    wb.close()
    return result


def parse_csv(path: Path) -> dict:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = [dict(r) for r in reader]
    return {path.name: rows}


def main():
    if len(sys.argv) < 2:
        print("Usage: parse_prd.py <file> [<file> ...]", file=sys.stderr)
        sys.exit(1)

    combined = {}
    for arg in sys.argv[1:]:
        p = Path(arg)
        if not p.exists():
            print(f"File not found: {p}", file=sys.stderr)
            sys.exit(1)
        if p.suffix.lower() in (".xlsx", ".xls"):
            combined.update(parse_xlsx(p))
        elif p.suffix.lower() == ".csv":
            combined.update(parse_csv(p))
        else:
            print(f"Unsupported file type: {p.suffix}", file=sys.stderr)
            sys.exit(1)

    json.dump(combined, sys.stdout, indent=2, default=str)


if __name__ == "__main__":
    main()
